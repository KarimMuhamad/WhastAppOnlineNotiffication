from selenium import webdriver
import datetime
from selenium.webdriver.common.by import By
import time
import pygetwindow as gw
import openpyxl
from  plyer import notification as ntf

inp = input('Contact Name : ')

ExcelFile = 'Data/Data.xlsx'

wb = openpyxl.load_workbook(ExcelFile)
ws = wb['Data']
rowN = ws.max_row

option = webdriver.ChromeOptions()
option.add_argument('--user-data-dir=C:\\Users\\---Use Yours Users Pc Name---\\AppData\\Local\\Google\\Chrome\\User Data\\Default')
option.add_argument('--profile-directory=Default')

ChromeDriverManager = 'chromedriver.exe'

driver = webdriver.Chrome(options=option)
driver.get('https://web.whatsapp.com/')

win = gw.getActiveWindow()

win.minimize()

def cek(u):
    if(u == []):
        return ""
    
    if(len(u)>= 0):
        d = u[0].text
        return d


ldate = datetime.datetime.now()
date = ldate.strftime('%Y-%m-%d')

time.sleep(15)

user = driver.find_element_by_xpath('//span[@title="{}"]'.format(inp))
user.click()

print("Start Tracking -> 😁")

i = 1

while True:

    t = driver.find_elements(By.XPATH, '/html/body/div[1]/div/div/div[4]/div/header/div[2]/div[2]/span')

    d = cek(t)
    
    time_off = None
    duration = 0
    time_on = None
    
    if d == "online" :
        time_on = datetime.datetime.now().strftime('%X')

        ntf.notify(
            title = inp,
            message = f'{time_on} Online',
            app_name = 'WhatsApp',
            timeout = 1
        )

        while True:
            t = driver.find_elements(By.XPATH, '/html/body/div[1]/div/div/div[4]/div/header/div[2]/div[2]/span')
            d = cek(t)

            if d != "online":
                break

            duration += 1
            time.sleep(1)
        
        time_off = datetime.datetime.now().strftime('%X')

        ws.cell(row=rowN+i, column=1, value=inp)
        ws.cell(row=rowN+i, column=2, value=date)
        ws.cell(row=rowN+i, column=3, value=time_on)
        ws.cell(row=rowN+i, column=4, value=time.strftime("%X", time.gmtime(duration)))
        ws.cell(row=rowN+i, column=5, value=time_off)
        wb.save(ExcelFile)
        print(f'[{i-1}] Saved')
        print(f'-> {duration}s {time_off}')

        i += 1

        ntf.notify(
            title = inp,
            message = f'{time_off} Offline',
            app_name = 'WhatsApp',
            timeout = 1
        )
